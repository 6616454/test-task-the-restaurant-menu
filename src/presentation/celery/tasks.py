from datetime import datetime, timedelta

from openpyxl.styles import Font
from openpyxl.workbook import Workbook


def collect_menu_data(report_menus: list[dict]) -> str:
    date = (datetime.now() + timedelta(hours=3)).strftime("%H:%M-%d.%m.%Y")
    filename = f"{date}_menu.xlsx"

    book = Workbook()
    sheet = book.active

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 35
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 35
    sheet.column_dimensions["F"].width = 10

    font = Font(name="Montserrat", bold=True)

    menu_row, menu_column, menu_counter = 1, 1, 1
    for menu in report_menus:
        menu_cells = [
            sheet.cell(menu_row, menu_column, value=menu_counter),
            sheet.cell(menu_row, menu_column + 1, value=menu["title"]),
            sheet.cell(menu_row, menu_column + 2, value=menu["description"]),
        ]

        for menu_cell in menu_cells:
            menu_cell.font = font

        menu_counter += 1

        submenu_row = menu_row + 1
        submenu_column = menu_column + 1

        submenu_counter = 1

        for submenu in menu["submenus"]:
            submenu_cells = [
                sheet.cell(submenu_row, submenu_column, value=submenu_counter),
                sheet.cell(submenu_row, submenu_column + 1, value=submenu["title"]),
                sheet.cell(
                    submenu_row, submenu_column + 2, value=submenu["description"]
                ),
            ]

            for submenu_cell in submenu_cells:
                submenu_cell.font = font

            submenu_counter += 1

            dish_row = submenu_row + 1
            dish_column = submenu_column + 1

            dish_counter = 1

            for dish in submenu["dishes"]:
                dish_cells = [
                    sheet.cell(dish_row, dish_column, value=dish_counter),
                    sheet.cell(dish_row, dish_column + 1, value=dish["title"]),
                    sheet.cell(dish_row, dish_column + 2, value=dish["description"]),
                    sheet.cell(dish_row, dish_column + 3, value=dish["price"]),
                ]

                for dish_cell in dish_cells:
                    dish_cell.font = font

                dish_counter += 1

                dish_row += 1

            submenu_row = dish_row

        menu_row = submenu_row

    book.save(f"data/{filename}")

    return f"http://127.0.0.1:8000/api/v1/report/download/{filename}"
